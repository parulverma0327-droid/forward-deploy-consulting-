#!/usr/bin/env python3
"""Generate UC2 Business glossary Excel — Location, Inventory Position, IP output."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "Business glossary.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HUB_HEADERS = [
    "Section",
    "Field Name",
    "Description",
    "Source Table(s)",
    "Derivation / Transformation",
    "Data Type",
    "Sample Value",
    "Notes",
]

AGENT_HEADERS = [
    "Section",
    "Field Name",
    "Description",
    "Populated by",
    "From (at run)",
    "Data Type",
    "Sample Value",
    "Notes",
]


def style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def auto_width(ws, max_width: int = 52) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


def write_rows(ws, headers: list[str], rows: list[tuple], start_row: int = 1) -> None:
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))
    for i, row in enumerate(rows, start_row + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    auto_width(ws)


def main() -> None:
    wb = Workbook()

    # --- README ---
    ws = wb.active
    ws.title = "README"
    readme = [
        ["UC2 Business Glossary — Inventory Planning", ""],
        ["Forward Deploy Consulting · Jul 2026", ""],
        ["", ""],
        ["Purpose", "Gold-layer data contract for store/DC inventory position and Inventory Policy outputs."],
        ["Audience", "Data engineer, UC2 build; UC1 agents that read inventory inputs (InventoryPolicyAgent handoff)."],
        ["Status", "TO-BE — synthetic demo data until client WMS feed is available."],
        ["", ""],
        ["Sheets", ""],
        ["Location", "Store and DC master — one row per physical location. Grain: location_id."],
        ["Inventory Position", "WMS on-hand snapshot per SKU × location. Grain: location_id + sku_id + snapshot_dt."],
        ["InventoryPolicyAgent", "Agent output — inventory_recommendation. Grain: sku_id + location_id + run_id."],
        ["UC1 Cross-Reference", "Which UC1 agent tables read from this workbook."],
        ["", ""],
        ["UC1 dependency", "UC1 InventoryPolicyAgent reads Location + Inventory Position from this glossary (not UC1 workbook)."],
        ["UC1 file", "use-cases/uc1-consumer-planning/glossary/Business glossary.xlsx — keeps demand_forecast_recommendation, replacement_score, agent loop tables."],
    ]
    for i, (a, b) in enumerate(readme, 1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=1).font = Font(bold=a in {"Purpose", "Audience", "Status", "Sheets", "UC1 dependency", "UC1 file"})
    auto_width(ws)

    # --- UC1 Cross-Reference ---
    ws = wb.create_sheet("UC1 Cross-Reference")
    write_rows(
        ws,
        ["UC1 table / agent", "UC1 field or read", "UC2 source sheet", "UC2 field", "Notes"],
        [
            (
                "InventoryPolicyAgent (UC1)",
                "Reads at run — location master",
                "Location",
                "location_id, location_type, region_canonical, parent_dc_id",
                "Expands demand_forecast_recommendation (region grain) to sku × location.",
            ),
            (
                "InventoryPolicyAgent (UC1)",
                "Reads at run — on-hand snapshot",
                "Inventory Position",
                "location_id, sku_id, units_on_hand, lead_time_days",
                "Compared to safety_stock_units to set exception_priority_flag and reorder_qty.",
            ),
            (
                "InventoryPolicyAgent (UC1)",
                "Writes inventory_recommendation",
                "InventoryPolicyAgent",
                "All WRITES columns",
                "Canonical field defs live in UC2; UC1 IP tab should reference this workbook.",
            ),
            (
                "TraceabilityAgent (UC1)",
                "Reads inventory_recommendation",
                "InventoryPolicyAgent",
                "safety_stock_units, exception_priority_flag",
                "Optional UC2 signal in recommendation envelope.",
            ),
            (
                "PublisherAgent (UC1)",
                "Reads inventory_recommendation (optional)",
                "InventoryPolicyAgent",
                "exception_priority_flag",
                "UC2 handoff — not shown on forecast grid.",
            ),
        ],
    )

    # --- Location ---
    ws = wb.create_sheet("Location")
    write_rows(
        ws,
        HUB_HEADERS,
        [
            (
                None,
                "— ENTITY: Location (Gold) —",
                "Canonical store and DC master. One row per physical location where inventory is held or sold. Used by Inventory Policy to expand regional demand to sku × location.",
                "—",
                "—",
                "—",
                "Grain: location_id",
                "[TO-BE] Source: WMS / store master / ERP. Demo: 2 stores + 1 DC in US-PNW.",
            ),
            (
                "Keys",
                "location_id",
                "Primary key for the location. Stable ID used across WMS, allocation, and inventory_recommendation. Same value on Inventory Position and inventory_recommendation.",
                "WMS; store master",
                "Pass-through from client location master",
                "string",
                "STORE-102",
                "PK. Format: STORE-{n} for stores, DC-{region}-{n} for DCs.",
            ),
            (
                None,
                "location_type",
                "Whether this location is a customer-facing store or a distribution center. Inventory Policy applies different safety-stock and reorder rules per type.",
                "WMS; store master",
                "Pass-through; enum store | dc",
                "string",
                "store",
                "store = sell-through / stockout risk. dc = regional buffer / replenishment source.",
            ),
            (
                None,
                "location_nm",
                "Human-readable location name for planner and ops UI.",
                "WMS; store master",
                "Pass-through",
                "string",
                "Seattle Flagship",
                "Not used in IP calculations.",
            ),
            (
                "Geo",
                "region_canonical",
                "Planning region this location belongs to (e.g. US-PNW). Join key to UC1 demand_forecast_recommendation.region_canonical when expanding region demand to locations.",
                "Geo Region; Finance / Planning master",
                "Map store/DC region code → region_canonical",
                "string",
                "US-PNW",
                "FK concept to UC1 Geo Region.region_canonical.",
            ),
            (
                None,
                "enterprise_region",
                "Client-specific region label as used in OMS / store master (e.g. Pacific NW). Optional alias for joins to UC1 hubs that use enterprise_region.",
                "store master; OMS",
                "Pass-through",
                "string",
                "Pacific NW",
                "Alias only — planning grain uses region_canonical.",
            ),
            (
                "Hierarchy",
                "parent_dc_id",
                "DC that replenishes this store. NULL for DC rows. Stores inherit regional buffer from parent DC when computing reorder from DC to store.",
                "Location",
                "FK → Location.location_id WHERE location_type = dc",
                "string",
                "DC-PNW-01",
                "NULL on DC rows. Required on store rows for demo.",
            ),
            (
                "Status",
                "is_active_ind",
                "Whether this location is active for inventory planning. Inactive locations are excluded from Inventory Policy scope.",
                "WMS; store master",
                "Pass-through; Y / N",
                "string",
                "Y",
                "Filter is_active_ind = Y at IP run.",
            ),
            (
                "Pipeline",
                "last_updated_dt",
                "UTC timestamp when this location row was last refreshed from source.",
                "Location pipeline",
                "Pipeline run timestamp",
                "timestamp",
                "2026-07-07 06:00:00",
                "[Metadata]",
            ),
        ],
    )

    # --- Inventory Position ---
    ws = wb.create_sheet("Inventory Position")
    write_rows(
        ws,
        HUB_HEADERS,
        [
            (
                None,
                "— ENTITY: Inventory Position (Gold) —",
                "Point-in-time inventory snapshot from WMS. One row per SKU × location per snapshot. Answers: how many units are on hand (and in transit) right now? Inventory Policy reads this — does not compute on-hand.",
                "—",
                "—",
                "—",
                "Grain: location_id + sku_id + snapshot_dt",
                "[TO-BE] Source: WMS / ERP inventory feed. Demo: static seed for hero SKU at 3 PNW locations.",
            ),
            (
                "Keys",
                "inventory_snapshot_id",
                "Unique row identifier for this snapshot record. System-generated on ingest.",
                "Inventory pipeline",
                "SYSTEM_UUID() on ingest",
                "string",
                "invsnap_001",
                "PK.",
            ),
            (
                None,
                "location_id",
                "Where stock is held. FK to Location.location_id (store or DC).",
                "Location",
                "Pass-through from WMS; validated against Location hub",
                "string",
                "STORE-102",
                "FK → Location.location_id.",
            ),
            (
                None,
                "sku_id",
                "SKU buying unit. FK to Product / UC1 demand_forecast_recommendation.sku_id.",
                "WMS; Product",
                "Pass-through",
                "string",
                "SKU-1025189-BLK-10",
                "Join to UC1 demand_forecast_recommendation on sku_id.",
            ),
            (
                None,
                "product_cd",
                "Style + color code for the SKU. Pass-through for joins to UC1 product_cd grain when sku_id is not on forecast row.",
                "Product",
                "Pass-through from Product hub",
                "string",
                "1025189-BLK",
                "Optional join key to UC1 product_cd.",
            ),
            (
                "Quantities",
                "units_on_hand",
                "Physical units available at this location at snapshot time. Sellable or allocatable stock — not on order, not in transit. Inventory Policy compares this to safety_stock_units to flag exceptions.",
                "WMS",
                "Pass-through from WMS inventory count",
                "integer",
                "18",
                "Core input for exception_priority_flag and reorder_qty.",
            ),
            (
                None,
                "units_in_transit",
                "Units shipped from origin but not yet received at this location. Optional — used when reorder logic accounts for inbound pipeline.",
                "WMS",
                "Pass-through from WMS in-transit status",
                "integer",
                "12",
                "Optional for v1 demo; NULL allowed.",
            ),
            (
                "Policy inputs",
                "lead_time_days",
                "Expected days from reorder decision to receipt at this location. Shorter for store (from DC), longer for DC (from vendor). Used in safety stock and reorder calculations.",
                "WMS; allocation config",
                "Pass-through or default by location_type (store=2, dc=14 demo)",
                "integer",
                "2",
                "Demo defaults: store=2, dc=14.",
            ),
            (
                "Snapshot",
                "snapshot_dt",
                "Date (or timestamp) when WMS took this inventory count. Freshness gate for Inventory Policy — stale snapshots may skip or flag the run.",
                "WMS",
                "Pass-through from WMS snapshot timestamp",
                "timestamp",
                "2026-07-07 06:00:00",
                "Use latest snapshot_dt per location_id + sku_id at IP run.",
            ),
            (
                "Pipeline",
                "last_updated_dt",
                "UTC timestamp when this row was last loaded into Gold.",
                "Inventory pipeline",
                "Pipeline run timestamp",
                "timestamp",
                "2026-07-07 06:15:00",
                "[Metadata]",
            ),
        ],
    )

    # --- InventoryPolicyAgent output ---
    ws = wb.create_sheet("InventoryPolicyAgent")
    write_rows(
        ws,
        AGENT_HEADERS,
        [
            (
                None,
                "— AGENT: InventoryPolicyAgent —",
                "Applies business rules after UC1 Demand Sensing and Replacement Cycle complete. Combines regional demand signals with store/DC on-hand to produce safety stock, exception flags, and reorder quantities. UC2 scope — not on UC1 forecast grid.",
                "—",
                "—",
                "—",
                "—",
                "[TO-BE] Runs in UC1 hourly loop as handoff agent; field definitions owned here.",
            ),
            (
                "Reads at run",
                "— inputs —",
                "UC1 agent_run_history (scope); UC1 demand_forecast_recommendation (adjusted demand); UC1 replacement_score (optional context); Location (store/DC master); Inventory Position (units_on_hand, lead_time_days).",
                "InventoryPolicyAgent (reads)",
                "See Notes",
                "—",
                "—",
                "Not columns on output table. UC1 forecast tables documented in UC1 Business glossary.xlsx.",
            ),
            (
                None,
                "— WRITES: inventory_recommendation —",
                "Safety stock and reorder recommendation per SKU × location per agent run.",
                "—",
                "—",
                "—",
                "Grain: sku_id + location_id + run_id",
                "One row per grain per run_id. Demo: 3 rows (2 stores + 1 DC) per hero SKU.",
            ),
            (
                "Keys",
                "inventory_rec_id",
                "Unique row identifier for this inventory recommendation (SYSTEM_UUID()). Primary key.",
                "InventoryPolicyAgent",
                "Runtime UUID",
                "string",
                "inv_001",
                "PK.",
            ),
            (
                None,
                "run_id",
                "Foreign key to UC1 agent_run_history — identifies which hourly agent loop produced this row.",
                "InventoryPolicyAgent",
                "agent_run_history",
                "string",
                "run_20260707_1400_a3f8",
                "FK → UC1 agent_run_history.run_id.",
            ),
            (
                None,
                "sku_id",
                "SKU this recommendation applies to. From UC1 Orchestrator scope / demand_forecast_recommendation.",
                "InventoryPolicyAgent",
                "demand_forecast_recommendation",
                "string",
                "SKU-1025189-BLK-10",
                "Same sku_id as UC1 forecast scope.",
            ),
            (
                None,
                "location_id",
                "Store or DC this recommendation applies to. From Location hub — one output row per active location in scope region.",
                "InventoryPolicyAgent",
                "Location",
                "string",
                "STORE-102",
                "FK → Location.location_id.",
            ),
            (
                "Policy",
                "safety_stock_units",
                "Recommended buffer stock for this SKU at this location. Derived from adjusted demand (units_intent_adjusted), confidence_score, lead_time_days, and location_type. Higher for DC than store at same demand.",
                "InventoryPolicyAgent",
                "SafetyStockSubAgent",
                "integer",
                "48",
                "Demo heuristic: f(units_intent_adjusted, confidence, lead_time, location_type). Formula TBD in agent spec.",
            ),
            (
                None,
                "exception_priority_flag",
                "TRUE when planner should review this location — e.g. on-hand below safety stock, demand spike vs baseline, or DC cannot cover store deficit. FALSE = within policy tolerance.",
                "InventoryPolicyAgent",
                "ExceptionPrioritySubAgent",
                "boolean",
                "TRUE",
                "TRUE when units_on_hand < safety_stock_units OR regional demand uplift exceeds threshold.",
            ),
            (
                None,
                "reorder_qty",
                "Suggested reorder quantity for this SKU at this location. Store: replenishment from parent DC. DC: replenishment from vendor. Zero when on-hand covers safety stock plus horizon demand.",
                "InventoryPolicyAgent",
                "Policy rules",
                "integer",
                "120",
                "Demo: MAX(0, safety_stock_units + horizon_demand − units_on_hand − units_in_transit).",
            ),
            (
                None,
                "policy_rule_version",
                "Version of the inventory policy rule set applied. Audit trail for which business rules produced this recommendation.",
                "InventoryPolicyAgent",
                "Policy config",
                "string",
                "inv_policy_v1",
                "[EXTERNAL — Policy config]",
            ),
            (
                None,
                "inputs_used",
                "JSON audit of key inputs that drove this row — e.g. units_intent_adjusted, units_on_hand, lead_time_days, location_type. For traceability and planner dispute resolution.",
                "InventoryPolicyAgent",
                "Runtime",
                "json",
                '{"units_intent_adjusted":332,"units_on_hand":18,"lead_time_days":2}',
                "Not shown on forecast grid; consumed by TraceabilityAgent.",
            ),
            (
                "Pipeline",
                "last_updated_dt",
                "UTC timestamp when InventoryPolicyAgent finished and wrote this row.",
                "InventoryPolicyAgent",
                "Runtime",
                "timestamp",
                "2026-07-07 14:01:45",
                None,
            ),
        ],
    )

    # --- Demo seed reference (not a hub — quick lookup) ---
    ws = wb.create_sheet("Demo Seed")
    write_rows(
        ws,
        ["Entity", "location_id / sku_id", "Field", "Value", "Story"],
        [
            ("Location", "STORE-102", "location_type", "store", "Low on-hand → exception"),
            ("Location", "STORE-118", "location_type", "store", "Healthy on-hand → no reorder"),
            ("Location", "DC-PNW-01", "location_type", "dc", "Regional buffer"),
            ("Inventory Position", "STORE-102", "units_on_hand", "18", "Below safety stock"),
            ("Inventory Position", "STORE-118", "units_on_hand", "95", "Above safety stock"),
            ("Inventory Position", "DC-PNW-01", "units_on_hand", "420", "Bulk regional stock"),
            ("inventory_recommendation", "STORE-102", "reorder_qty", "120", "Exception row"),
            ("inventory_recommendation", "STORE-118", "reorder_qty", "0", "No action"),
            ("inventory_recommendation", "DC-PNW-01", "reorder_qty", "340", "DC replenishment"),
        ],
    )

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
