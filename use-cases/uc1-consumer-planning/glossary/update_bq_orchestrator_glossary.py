#!/usr/bin/env python3
"""Add ForecastSignalBQAgent + signal_delta sheets; refresh OrchestratorAgent output."""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GLOSSARY = Path(__file__).resolve().parent / "Business glossary.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

AGENT_HEADERS = [
    "Section",
    "Field Name",
    "Description",
    "Populated by",
    "From (at run)",
    "Data Type",
    "Sample Value",
    "Notes",
]

HUB_HEADERS = [
    "Section",
    "Field Name",
    "Description",
    "Source Table(s)",
    "Derivation / Transformation",
    "Data Type",
    "Sample Value",
    "Notes",
]


def style_header_row(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def auto_width(ws, max_width: int = 52) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


def write_sheet(ws, headers: list[str], rows: list[tuple]) -> None:
    ws.delete_rows(1, ws.max_row)
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header_row(ws, 1, len(headers))
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    auto_width(ws)


def main() -> None:
    wb = load_workbook(GLOSSARY)

    # --- ForecastSignalBQAgent ---
    if "ForecastSignalBQAgent" in wb.sheetnames:
        del wb["ForecastSignalBQAgent"]
    ws = wb.create_sheet("ForecastSignalBQAgent", 0)
    write_sheet(
        ws,
        AGENT_HEADERS,
        [
            (
                None,
                "— AGENT: ForecastSignalBQAgent —",
                "BigQuery-native agent on Gold. Detects signal changes since last run, writes signal_delta. Same agent serves Inquiry on-demand (NL→SQL). Does not route specialists or score unit lifts.",
                "—",
                "—",
                "—",
                "—",
                "[TO-BE] Build on demandsensinglayer.dsl_dataset.",
            ),
            (
                "Reads at run",
                "— inputs —",
                "Clickstream Agg; DemandForecast_AsIS; Member Hub; agent_run_history (last run boundary); Geo Region.",
                "ForecastSignalBQAgent (reads)",
                "See Notes",
                "—",
                "—",
                "Not columns on output table.",
            ),
            (
                None,
                "— WRITES: signal_delta —",
                "One row per detected scope change per run cycle.",
                "—",
                "—",
                "—",
                "Grain: product_cd + region + week + delta_type",
                "OrchestratorAgent reads this table — no manual scope bypass.",
            ),
            (
                "Keys",
                "signal_delta_id",
                "Unique row identifier for this delta record (SYSTEM_UUID()). Primary key.",
                "ForecastSignalBQAgent",
                "Runtime UUID",
                "string",
                "sd_001",
                "PK.",
            ),
            (
                None,
                "product_cd",
                "Style + color where the consumer or forecast signal moved. Join key to DS/RC scope.",
                "ForecastSignalBQAgent",
                "Clickstream Agg or DemandForecast_AsIS",
                "string",
                "1025189-BLK",
                None,
            ),
            (
                None,
                "region",
                "Enterprise region label from the source hub row (e.g. Pacific NW).",
                "ForecastSignalBQAgent",
                "Clickstream Agg.region",
                "string",
                "Pacific NW",
                "Mapped to region_canonical via Geo Region.",
            ),
            (
                None,
                "region_canonical",
                "Planning region after geo map (e.g. US-PNW). Orchestrator uses this in scope_json.",
                "ForecastSignalBQAgent",
                "Geo Region",
                "string",
                "US-PNW",
                "FK concept → Geo Region.region_canonical.",
            ),
            (
                None,
                "week",
                "Forecast week label where the change applies (e.g. 2026-W45).",
                "ForecastSignalBQAgent",
                "Clickstream Agg or DemandForecast_AsIS",
                "string",
                "2026-W45",
                None,
            ),
            (
                "Delta",
                "delta_type",
                "Classification of what changed: intent_spike, intent_drop, baseline_shift, member_model_refresh.",
                "ForecastSignalBQAgent",
                "Runtime rules",
                "string",
                "intent_spike",
                "Demo thresholds: intent ≥15% change; baseline ≥10% change.",
            ),
            (
                None,
                "delta_magnitude",
                "Size of the detected change — percent or absolute difference in the triggering metric.",
                "ForecastSignalBQAgent",
                "Formula",
                "decimal",
                "0.22",
                "ABS(current − prior) / prior for intent; same pattern for baseline.",
            ),
            (
                None,
                "source_table",
                "Glossary hub that triggered this row — audit for traceability.",
                "ForecastSignalBQAgent",
                "Runtime",
                "string",
                "Clickstream Agg",
                "Clickstream Agg | DemandForecast_AsIS | Member Hub.",
            ),
            (
                None,
                "prior_value",
                "Metric value at last completed agent run (or lookback snapshot). Audit only.",
                "ForecastSignalBQAgent",
                "Prior snapshot",
                "decimal",
                "820.0",
                "weighted_intent_score or units_forecast depending on delta_type.",
            ),
            (
                None,
                "current_value",
                "Metric value at current detection time. Audit only.",
                "ForecastSignalBQAgent",
                "Current Gold row",
                "decimal",
                "1000.0",
                None,
            ),
            (
                "Pipeline",
                "detected_at",
                "UTC timestamp when ForecastSignalBQAgent wrote this delta row.",
                "ForecastSignalBQAgent",
                "Runtime",
                "timestamp",
                "2026-07-07 13:55:00",
                "Orchestrator reads rows where detected_at > last completed_at.",
            ),
        ],
    )

    # --- signal_delta (entity reference — same fields, hub-style for validation) ---
    if "signal_delta" in wb.sheetnames:
        del wb["signal_delta"]
    ws = wb.create_sheet("signal_delta")
    write_sheet(
        ws,
        HUB_HEADERS,
        [
            (
                None,
                "— ENTITY: signal_delta (Gold) —",
                "Routing table written by ForecastSignalBQAgent. Lists SKU × region × week scopes where signals changed since the last agent run. OrchestratorAgent reads — does not write.",
                "—",
                "—",
                "—",
                "Grain: product_cd + region + week + delta_type",
                "[TO-BE] Populated hourly by BQ agent before Orchestrator runs.",
            ),
            (
                "Keys",
                "signal_delta_id",
                "Primary key.",
                "ForecastSignalBQAgent",
                "SYSTEM_UUID()",
                "string",
                "sd_001",
                "PK.",
            ),
            (
                None,
                "product_cd",
                "Style + color in scope.",
                "ForecastSignalBQAgent",
                "Clickstream Agg; DemandForecast_AsIS",
                "string",
                "1025189-BLK",
                None,
            ),
            (
                None,
                "region",
                "Enterprise region from source.",
                "ForecastSignalBQAgent",
                "Clickstream Agg",
                "string",
                "Pacific NW",
                None,
            ),
            (
                None,
                "region_canonical",
                "Planning region.",
                "ForecastSignalBQAgent",
                "Geo Region",
                "string",
                "US-PNW",
                None,
            ),
            (
                None,
                "week",
                "Affected week.",
                "ForecastSignalBQAgent",
                "Clickstream Agg; DemandForecast_AsIS",
                "string",
                "2026-W45",
                None,
            ),
            (
                None,
                "delta_type",
                "Change class.",
                "ForecastSignalBQAgent",
                "Runtime rules",
                "string",
                "intent_spike",
                None,
            ),
            (
                None,
                "delta_magnitude",
                "Change size.",
                "ForecastSignalBQAgent",
                "Formula",
                "decimal",
                "0.22",
                None,
            ),
            (
                None,
                "source_table",
                "Triggering hub.",
                "ForecastSignalBQAgent",
                "Runtime",
                "string",
                "Clickstream Agg",
                None,
            ),
            (
                None,
                "prior_value",
                "Before metric.",
                "ForecastSignalBQAgent",
                "Prior snapshot",
                "decimal",
                "820.0",
                None,
            ),
            (
                None,
                "current_value",
                "After metric.",
                "ForecastSignalBQAgent",
                "Current Gold row",
                "decimal",
                "1000.0",
                None,
            ),
            (
                "Pipeline",
                "detected_at",
                "When detected.",
                "ForecastSignalBQAgent",
                "Runtime",
                "timestamp",
                "2026-07-07 13:55:00",
                None,
            ),
        ],
    )

    # --- OrchestratorAgent output (refresh) ---
    ws = wb["OrchestratorAgent output"]
    write_sheet(
        ws,
        AGENT_HEADERS,
        [
            (
                None,
                "— AGENT: OrchestratorAgent —",
                "Hourly traffic controller. Reads signal_delta from ForecastSignalBQAgent, runs gates, builds scope_json, issues run_id, routes RC ∥ DS → IP → Traceability → Publisher. Does not detect deltas or score forecasts.",
                "—",
                "—",
                "—",
                "—",
                "[TO-BE]",
            ),
            (
                "Reads at run",
                "— inputs —",
                "signal_delta (from BQ agent); Clickstream Agg (freshness); DemandForecast_AsIS (freshness); Member Hub (model refresh).",
                "OrchestratorAgent (reads)",
                "See Notes",
                "—",
                "—",
                "Not columns on output table. No UI/manual scope bypass.",
            ),
            (
                None,
                "— WRITES: agent_run_history —",
                "One row per hourly agent loop.",
                "—",
                "—",
                "—",
                "Grain: run_id",
                "One row per grain per run_id.",
            ),
            (
                "Keys",
                "run_id",
                "Primary key for this hourly loop (run_{YYYYMMDD_HHMM}_{uuid8}). Foreign key on all downstream agent output tables.",
                "OrchestratorAgent",
                "Runtime UUID",
                "string",
                "run_20260707_1400_a3f8",
                "FK for all agent tables.",
            ),
            (
                None,
                "run_status",
                "Outcome of the loop: completed (specialists ran), skipped (gates failed or no delta), failed (error mid-run).",
                "OrchestratorAgent",
                "Runtime",
                "enum",
                "completed",
                "completed | skipped | failed",
            ),
            (
                None,
                "gates_passed",
                "TRUE if all Bucket 1 freshness checks passed. FALSE → DS/RC/IP do not run.",
                "OrchestratorAgent",
                "Runtime",
                "boolean",
                "TRUE",
                "Clickstream + forecast freshness < 2h; member model current.",
            ),
            (
                None,
                "scope_json",
                "JSON list of routed scopes from signal_delta: product_cd, sku_id, region_canonical, week. DS + RC read this.",
                "OrchestratorAgent",
                "signal_delta + Product",
                "json",
                '[{"product_cd":"1025189-BLK","sku_id":"SKU-1025189-BLK-10","region_canonical":"US-PNW","week":"2026-W45"}]',
                "Built from signal_delta rows after dedup.",
            ),
            (
                None,
                "agents_executed",
                "JSON list of agents that ran this loop — audit trail.",
                "OrchestratorAgent",
                "Runtime",
                "json",
                '["BQ","ORCH","RC","DS","IP","TR","PUB"]',
                None,
            ),
            (
                "Time",
                "started_at",
                "UTC timestamp when Orchestrator began this run (after BQ agent wrote signal_delta).",
                "OrchestratorAgent",
                "Runtime",
                "timestamp",
                "2026-07-07 14:00:00",
                None,
            ),
            (
                None,
                "completed_at",
                "UTC timestamp when Orchestrator finished (after Publisher or skip).",
                "OrchestratorAgent",
                "Runtime",
                "timestamp",
                "2026-07-07 14:02:15",
                None,
            ),
        ],
    )

    wb.save(GLOSSARY)
    print(f"Updated {GLOSSARY}")


if __name__ == "__main__":
    main()
