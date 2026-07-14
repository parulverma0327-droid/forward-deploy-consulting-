#!/usr/bin/env python3
"""Generate Show & Tell planner grid business glossary Excel."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "Show and Tell — Planner Grid Glossary.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER


def auto_width(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def write_table(ws, headers, rows, start_row=1):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))
    for i, row in enumerate(rows, start_row + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    auto_width(ws)
    return start_row + len(rows) + 2


WEEKS = [
    ("W42", "2026-10-13", "2026-10-19", 240, 310, 298),
    ("W43", "2026-10-20", "2026-10-26", 265, 340, None),
    ("W44", "2026-10-27", "2026-11-02", 270, 355, None),
    ("W45", "2026-11-03", "2026-11-09", 280, 365, None),
    ("W46", "2026-11-10", "2026-11-16", 275, 360, None),
    ("W47", "2026-11-17", "2026-11-23", 260, 335, None),
    ("W48", "2026-11-24", "2026-11-30", 260, 335, None),
]

SKU = "KEEN-1024"
STYLE = "Targhee IV Mid WP · Earth Brown"
REGION = "US-PNW"
SEASON = "FA26"
CATEGORY = "Footwear > Trail"


def main():
    wb = Workbook()

    # --- README ---
    ws = wb.active
    ws.title = "README"
    readme = [
        ["Show & Tell — Planner Grid Glossary", ""],
        ["Forward Deploy Consulting · Jun 2026", ""],
        ["", ""],
        ["Purpose", "Data contract + sample grid for demand forecast UI and Planning Data Hub Gold"],
        ["Audience", "Data engineer, UI (Anil), planners, client glossary workshops"],
        ["Grain", "sku_id × region_canonical × week (date)"],
        ["", ""],
        ["Tabs", ""],
        ["Session_Fields", "Wizard inputs — season, category, region, from/to dates"],
        ["Forecast_Grid_Fields", "Dictionary of every field shown in the forecast grid"],
        ["Merch_Context_Fields", "Panel 1 read-only fields"],
        ["Insights_Fields", "Panel 3 signal explanations (units language)"],
        ["Agent_Glossary", "Agent + data-product terms; agreed vs blueprint status"],
        ["Agent_Status_Legend", "Definitions for status column on Agent_Glossary"],
        ["Sample_Long", "PDH-ready long format (engineer / pipeline)"],
        ["Sample_Wide_UI", "What the planner sees — weeks as columns (UI mock)"],
        ["Week_Calendar", "Week labels mapped to from/to dates for sample session"],
    ]
    for r, row in enumerate(readme, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws["A1"].font = Font(bold=True, size=14)
    auto_width(ws)

    # --- Session_Fields ---
    ws = wb.create_sheet("Session_Fields")
    headers = [
        "field_name",
        "ui_label",
        "required",
        "type",
        "source_hub",
        "entity",
        "example_value",
        "description",
    ]
    rows = [
        ["session_id", "Session ID", "Y", "UUID", "Planning Data Hub", "PlanningSession", "sess_fa26_trail_pnw_001", "Unique session key"],
        ["season_code", "Season", "Y", "string", "Product Hub / Season", "Season", "FA26", "Which commercial season — context for merch plan"],
        ["category", "Category", "Y", "string", "Product Hub", "Product", "Footwear > Trail", "Product hierarchy scope for grid rows"],
        ["region_canonical", "Region", "Y", "string", "Location Hub", "Location", "US-PNW", "Planning region; maps store clusters"],
        ["date_from", "From date", "Y", "date", "App input", "PlanningSession", "2026-10-13", "Start of forecast window (inclusive)"],
        ["date_to", "To date", "Y", "date", "App input", "PlanningSession", "2026-11-24", "End of forecast window (inclusive)"],
        ["consumer_signals_enabled", "Consumer signals", "N", "boolean", "App input", "PlanningSession", "FALSE / TRUE", "OFF=Run1 ON=Run2; rerun without changing dates"],
        ["session_status", "Status", "N", "enum", "Planning Data Hub", "PlanningSession", "draft | pending_review | approved", "Workflow state"],
        ["planner_id", "Planner", "Y", "string", "Auth", "User", "parvi.k", "Who owns the session"],
    ]
    write_table(ws, headers, rows)

    # --- Forecast_Grid_Fields ---
    ws = wb.create_sheet("Forecast_Grid_Fields")
    headers = [
        "field_name",
        "ui_label",
        "grid_axis",
        "type",
        "source_hub",
        "entity",
        "visible_run1",
        "visible_run2",
        "editable",
        "saved_to_pdh",
        "example_value",
        "description",
    ]
    rows = [
        ["sku_id", "SKU", "row", "string", "Product Hub", "Product", "Y", "Y", "N", "Y", "KEEN-1024", "Row key — buying unit"],
        ["style_name", "Style", "row", "string", "Product Hub", "Product", "Y", "Y", "N", "N", "Targhee IV Mid WP", "Display name"],
        ["region_canonical", "Region", "row/filter", "string", "Location Hub", "Location", "Y", "Y", "N", "Y", "US-PNW", "Session filter; repeated per row in long format"],
        ["week", "Week ending", "column", "date", "Derived", "DemandForecast", "Y", "Y", "N", "Y", "2026-10-19", "Column header = week end date; one column per week in date_from–date_to"],
        ["week_label", "Week", "column", "string", "Derived", "DemandForecast", "Y", "Y", "N", "N", "W42", "Display label (retailer fiscal week)"],
        ["units_historical", "Baseline forecast", "cell", "integer", "Planning Data Hub", "DemandForecast", "Y", "Y", "N", "Y", "240", "Run 1 — sales history + promo calendar"],
        ["units_intent_adjusted", "Enriched forecast", "cell", "integer", "Planning Data Hub", "DemandForecast", "N", "Y", "N", "Y", "310", "Run 2 — baseline + Forward Demand Signal"],
        ["units_delta", "Δ vs baseline", "cell", "integer", "Derived", "DemandForecast", "N", "Y", "N", "N", "+70", "units_intent_adjusted − units_historical; UI only"],
        ["units_actual", "Actuals", "cell", "integer", "OMS / Sales Hub", "Transaction", "Y", "Y", "N", "N", "298", "Sold units in week where known; blank if future"],
        ["units_override", "Override", "cell", "integer", "App input", "DemandForecast", "Y", "Y", "Y", "Y", "—", "Planner manual edit; wins over system if set"],
        ["units_final", "Approved forecast", "cell", "integer", "Planning Data Hub", "DemandForecast", "Y", "Y", "N", "Y", "310", "Saved on Accept — downstream handoff"],
        ["confidence_score", "Confidence", "row/summary", "float", "Planning Data Hub", "DemandForecast", "Y", "Y", "N", "Y", "0.82", "Model confidence 0–1"],
        ["consumer_signals_applied", "Signals applied", "session", "boolean", "Planning Data Hub", "DemandForecast", "Y", "Y", "N", "Y", "TRUE", "Audit: which run produced units_final"],
        ["row_total_baseline", "Total baseline", "row summary", "integer", "Derived", "DemandForecast", "Y", "Y", "N", "N", "1850", "Sum units_historical across weeks"],
        ["row_total_enriched", "Total enriched", "row summary", "integer", "Derived", "DemandForecast", "N", "Y", "N", "N", "2400", "Sum units_intent_adjusted across weeks"],
    ]
    write_table(ws, headers, rows)

    # --- Merch_Context_Fields ---
    ws = wb.create_sheet("Merch_Context_Fields")
    headers = ["field_name", "ui_label", "type", "source_hub", "example_value", "description"]
    rows = [
        ["season_code", "Season", "string", "Merch slice", "FA26", "Read-only context"],
        ["category", "Category", "string", "Merch slice", "Footwear > Trail", "Matches session"],
        ["otb_remaining_usd", "OTB remaining", "decimal", "Merch slice", "4200000", "Open-to-buy dollars left"],
        ["sales_vs_plan_pct", "YTD sales vs plan", "decimal", "Merch slice", "1.02", "102%"],
        ["margin_vs_plan_pct", "YTD margin vs plan", "decimal", "Merch slice", "0.98", "98%"],
        ["planner_note", "Planner note", "string", "App / manual", "Trail running up; waterproof lagging PNW", "Optional free text"],
    ]
    write_table(ws, headers, rows)

    # --- Insights_Fields ---
    ws = wb.create_sheet("Insights_Fields")
    headers = [
        "insight_id",
        "ui_title",
        "source_hub",
        "source_entity",
        "impact_field",
        "status",
        "example_detail",
        "example_impact",
    ]
    rows = [
        ["INS-01", "Search spike", "Clickstream → PDH translate", "IntentSignal", "intent_units_lift", "demo_illustration", "+34% waterproof hiking boot PNW W42–44", "+220 units"],
        ["INS-02", "Browse-to-buy gap", "Clickstream + Product", "IntentSignal", "intent_units_lift", "demo_illustration", "High-CLV members viewed; size 10 OOS 6 days", "+180 units"],
        ["INS-03", "Cart abandon", "Clickstream → PDH translate", "IntentSignal", "flag only", "demo_illustration", "412 abandons; 61% size-related", "Sizing flag"],
        ["INS-04", "Member vs guest", "Member + OMS", "Transaction", "units_member / units_guest", "demo_illustration", "Member demand +31% vs baseline; guest flat", "Split in grid filter"],
        ["INS-05", "Replacement cadence", "OMS + Data Science", "DataScienceOutput", "units_replacement_adjusted", "demo_illustration", "Repeat buyers due W45–46", "+150 units"],
    ]
    write_table(ws, headers, rows)

    # --- Agent_Status_Legend ---
    ws = wb.create_sheet("Agent_Status_Legend")
    headers = ["status", "meaning", "safe_for_show_and_tell"]
    rows = [
        ["agreed_demo", "Signed off for Show & Tell — use in static PDH, UI, and talk track", "Y"],
        ["agreed_prod", "Real Phase-1/2 architecture — not built or wired in demo yet", "Y (describe concept only)"],
        ["blueprint_placeholder", "In Planner Loop PDF or draft spec — not discussed or validated in meetings", "N — do not cite as final logic"],
        ["demo_illustration", "Hero SKU sample numbers (Targhee IV) — illustrative only", "Y — label as example"],
    ]
    write_table(ws, headers, rows)

    # --- Agent_Glossary ---
    ws = wb.create_sheet("Agent_Glossary")
    headers = [
        "term",
        "layer",
        "definition",
        "grain",
        "status",
        "source",
        "pdh_field",
        "related_tab",
        "notes",
    ]
    rows = [
        [
            "OrchestratorAgent",
            "Agent",
            "Hourly traffic controller — checks data gates, reads signal deltas, routes work to specialist agents",
            "run (hourly)",
            "agreed_prod",
            "Planner Loop PDF §4; May 28 meeting",
            "—",
            "—",
            "Demo: skip or demo_mode; static POC uses UI toggle instead",
        ],
        [
            "DemandSensingAgent",
            "Agent",
            "Anchor agent — detects demand intent shifts 4–8 weeks ahead of sales; proposes unit adjustments on PDH",
            "sku_id × region × week",
            "agreed_demo",
            "Planner Loop PDF §4; UC1 doc; May 28 meeting",
            "units_intent_adjusted",
            "Forecast_Grid_Fields",
            "Recommend-only; planner approves → units_final",
        ],
        [
            "ClickstreamIntent (sub-agent)",
            "Sub-agent",
            "Scores browse, search, add-to-cart, wishlist, abandon from IntentEvent mart",
            "sku_id × region × week",
            "agreed_prod",
            "Planner Loop PDF §4",
            "intent_units_lift",
            "Insights_Fields",
            "Part of DemandSensingAgent",
        ],
        [
            "LoyaltySegment (sub-agent)",
            "Sub-agent",
            "Reads tier shifts, engagement, segment migration from CustomerSegment mart",
            "segment × geo × week",
            "agreed_prod",
            "Planner Loop PDF §4",
            "—",
            "—",
            "Part of DemandSensingAgent",
        ],
        [
            "CampaignResponse (sub-agent)",
            "Sub-agent",
            "Reads campaign impressions/clicks/conversions by segment × day",
            "campaign × segment × day",
            "agreed_prod",
            "Planner Loop PDF §4",
            "—",
            "—",
            "P1 priority in Phase-1 scope",
        ],
        [
            "Forward Demand Signal",
            "Data product",
            "Aggregated demand indicator from intent + replacement signals — what customers will buy, not have bought",
            "sku_id × region × week",
            "agreed_demo",
            "UC1 doc; Retail Ontology; Show & Tell spec",
            "units_intent_adjusted",
            "Forecast_Grid_Fields",
            "Lands in PDH via translate layer; feeds DemandSensingAgent",
        ],
        [
            "IntentSignal",
            "Entity",
            "Clickstream events (search, PDP, cart, wishlist) cleansed in Silver, aggregated for planning",
            "event_id (Silver); style × region × week (Gold)",
            "agreed_demo",
            "Planner Loop PDF §6; Show & Tell spec",
            "intent_units_lift",
            "Insights_Fields",
            "Never shown raw in UI — translate to units only",
        ],
        [
            "Weighted intent index",
            "Logic",
            "Combines event types into a single intent score before unit translation",
            "sku_id × region × week",
            "blueprint_placeholder",
            "Planner Loop PDF §4 DemandSensing Logic row",
            "intent_units_lift",
            "—",
            "PDF example weights only — tune per client; Basu did not specify in meeting",
        ],
        [
            "Intent event weights (browse/search/wishlist/cart)",
            "Logic",
            "browse(0.1), search(0.3), wishlist(0.5), add-to-cart(0.6) — relative weights in PDF",
            "per event type",
            "blueprint_placeholder",
            "Planner Loop PDF §4 only",
            "—",
            "—",
            "Do not present as validated; placeholder for BQML feature engineering",
        ],
        [
            "intent_units_lift",
            "PDH field",
            "Unit impact from translated intent signal on baseline forecast",
            "sku_id × region × week",
            "agreed_demo",
            "Show & Tell spec translate table",
            "intent_units_lift",
            "Insights_Fields",
            "Rolls into units_intent_adjusted",
        ],
        [
            "units_intent_adjusted",
            "PDH field",
            "Run 2 enriched forecast — baseline + Forward Demand Signal",
            "sku_id × region × week",
            "agreed_demo",
            "Show & Tell spec",
            "units_intent_adjusted",
            "Forecast_Grid_Fields",
            "Written by DemandSensingAgent; shown when consumer signals ON",
        ],
        [
            "units_historical",
            "PDH field",
            "Run 1 baseline — sales history + promo calendar only",
            "sku_id × region × week",
            "agreed_demo",
            "Show & Tell spec",
            "units_historical",
            "Forecast_Grid_Fields",
            "No agent adjustment; consumer_signals_applied = false",
        ],
        [
            "confidence_score",
            "PDH field",
            "Model confidence 0–1 on enriched forecast",
            "sku_id × region × week",
            "agreed_demo",
            "Planner Loop PDF; Show & Tell spec",
            "confidence_score",
            "Forecast_Grid_Fields",
            "Hero SKU example: 0.82",
        ],
        [
            "Signal delta",
            "Data artifact",
            "Scopes where consumer signals changed since last run — Orchestrator routing input",
            "sku_id × region × week",
            "agreed_prod",
            "Planner Loop PDF; May 28 meeting",
            "—",
            "—",
            "Demo: pre-stage one row per hero SKU",
        ],
        [
            "Recommendation (demand_uplift)",
            "Event / card",
            "Standardized recommendation emitted to Insight API / Pub/Sub for planner review",
            "recommendation_id",
            "agreed_demo",
            "Planner Loop PDF §7",
            "—",
            "Insights_Fields",
            "Status: pending → approved / rejected / overridden",
        ],
        [
            "Top drivers (feature contributions)",
            "Traceability",
            "SHAP-style weights explaining adjustment — e.g. clickstream_intent 45%",
            "per recommendation",
            "demo_illustration",
            "Planner Loop PDF JSON sample",
            "—",
            "Insights_Fields",
            "Use INS-01…05 in demo; % split is illustrative",
        ],
        [
            "BQML demand_sensing model",
            "Model",
            "BigQuery ML model scoring demand probability from intent features",
            "sku_id × region × week",
            "blueprint_placeholder",
            "Planner Loop PDF (demand_sensing_v3 example)",
            "—",
            "—",
            "Basu: one forecast + signal model on BQ — name/version TBD",
        ],
        [
            "4–8 week sensing horizon",
            "Concept",
            "How far ahead DemandSensing detects intent shifts before they appear in POS sales",
            "—",
            "agreed_prod",
            "Planner Loop PDF §4",
            "—",
            "—",
            "Concept agreed; exact window may vary by category",
        ],
        [
            "Hourly agent run",
            "Cadence",
            "Orchestrator + DemandSensing run on Cloud Scheduler / Pub/Sub",
            "hourly",
            "agreed_prod",
            "Planner Loop PDF §2.3; delivery plan Wk 7–8",
            "—",
            "—",
            "Not required for static Show & Tell POC",
        ],
        [
            "Client data-to-decision tuning",
            "Governance",
            "Per-client customization of which signals matter and how they map to planning actions",
            "—",
            "agreed_prod",
            "May 21 meeting (Basu)",
            "—",
            "—",
            "Replaces hard-coded intent weights in production",
        ],
        [
            "ReplacementCycleAgent",
            "Agent",
            "Models repurchase likelihood by segment × SKU — parallel to DemandSensing, not a substitute",
            "segment × sku_id",
            "agreed_prod",
            "Planner Loop PDF §4; May 28 meeting",
            "units_replacement_adjusted",
            "Insights_Fields (INS-05)",
            "Separate signal path; InventoryPolicy marries both",
        ],
        [
            "InventoryPolicyAgent",
            "Agent",
            "Translates demand + replacement signals into safety stock recommendations",
            "sku_id × location",
            "agreed_prod",
            "Planner Loop PDF §4",
            "—",
            "—",
            "Downstream of DemandSensingAgent",
        ],
        [
            "TraceabilityAgent",
            "Agent",
            "Attaches lineage package — model version, features, agent_reasoning — to every recommendation",
            "per recommendation",
            "agreed_prod",
            "Planner Loop PDF §4; May 28 meeting",
            "—",
            "Insights_Fields",
            "agent_reasoning = NL explanation for planner",
        ],
        [
            "PublisherAgent",
            "Agent",
            "Emits RecommendationEmitted to Pub/Sub + Insight API store",
            "per recommendation",
            "agreed_prod",
            "Planner Loop PDF §4",
            "—",
            "—",
            "Idempotent publish by recommendation_id",
        ],
        [
            "InquiryAgent",
            "Agent",
            "On-demand NL planner questions via Inquiry API — not part of hourly sensing loop",
            "per session",
            "agreed_prod",
            "Planner Loop PDF §4; §7",
            "—",
            "—",
            "SLA < 60s; separate from DemandSensing",
        ],
        [
            "Run 1",
            "Demo concept",
            "Forecast without consumer signals — units_historical only",
            "session",
            "agreed_demo",
            "Show & Tell spec",
            "units_historical",
            "Forecast_Grid_Fields",
            "consumer_signals_enabled = FALSE",
        ],
        [
            "Run 2",
            "Demo concept",
            "Same scope/dates with Forward Demand Signal applied — units_intent_adjusted",
            "session",
            "agreed_demo",
            "Show & Tell spec",
            "units_intent_adjusted",
            "Forecast_Grid_Fields",
            "Toggle ON → rerun; delta = sales story",
        ],
        [
            "Planning Data Hub (PDH)",
            "Hub",
            "Gold layer for planning — primary read/write for UI and agents at forecast grain",
            "sku_id × region × week",
            "agreed_demo",
            "Show & Tell spec",
            "DemandForecast entity",
            "Forecast_Grid_Fields",
            "Gold mart for planning; not all gold marts are PDH",
        ],
    ]
    write_table(ws, headers, rows)

    # --- Week_Calendar ---
    ws = wb.create_sheet("Week_Calendar")
    headers = ["week_label", "week_start_date", "week_end_date", "fiscal_week_num", "in_sample_session"]
    rows = [
        [w[0], w[1], w[2], int(w[0][1:]), "Y" if w[0] <= "W48" else "N"]
        for w in WEEKS
    ]
    write_table(ws, headers, rows)

    # --- Sample_Long ---
    ws = wb.create_sheet("Sample_Long")
    headers = [
        "session_id",
        "season_code",
        "category",
        "region_canonical",
        "sku_id",
        "style_name",
        "week_label",
        "week_end_date",
        "units_historical",
        "units_intent_adjusted",
        "units_delta",
        "units_actual",
        "units_override",
        "units_final",
        "confidence_score",
        "consumer_signals_applied",
    ]
    session = "sess_fa26_trail_pnw_001"
    long_rows = []
    for w in WEEKS:
        label, start, end, hist, enrich, actual = w
        delta = enrich - hist
        long_rows.append([
            session, SEASON, CATEGORY, REGION, SKU, STYLE,
            label, end, hist, enrich, delta,
            actual if actual is not None else "",
            "", enrich, 0.82, "TRUE",
        ])
    write_table(ws, headers, long_rows)

    # --- Sample_Wide_UI ---
    ws = wb.create_sheet("Sample_Wide_UI")
    r = 1
    ws.cell(row=r, column=1, value="SAMPLE — Planner grid (wide layout · one SKU · Run 2 visible)")
    ws.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 2

    meta = [
        ["Session", session],
        ["Season", SEASON],
        ["Category", CATEGORY],
        ["Region", REGION],
        ["SKU", SKU],
        ["Style", STYLE],
        ["From date", "2026-10-13"],
        ["To date", "2026-11-24"],
        ["Consumer signals", "ON (Run 2)"],
    ]
    for label, val in meta:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    r += 1

    # Column headers: Metric | W42 ... W48 | Total
    week_labels = [w[0] for w in WEEKS]
    week_ends = [w[2] for w in WEEKS]
    headers_wide = ["Metric (field_name)"] + week_labels + ["Row total"]
    for j, h in enumerate(headers_wide, 1):
        ws.cell(row=r, column=j, value=h)
    style_header_row(ws, r, len(headers_wide))
    r += 1
    # Sub-header row with week end dates
    ws.cell(row=r, column=1, value="Week ending (date)")
    for j, d in enumerate(week_ends, 2):
        ws.cell(row=r, column=j, value=d)
    for c in range(1, len(headers_wide) + 1):
        ws.cell(row=r, column=c).fill = SUBHEADER_FILL
        ws.cell(row=r, column=c).border = BORDER
    r += 1

    hist_vals = [w[3] for w in WEEKS]
    enrich_vals = [w[4] for w in WEEKS]
    actual_vals = [w[5] if w[5] is not None else "—" for w in WEEKS]
    delta_vals = [e - h for h, e in zip(hist_vals, enrich_vals)]

    metric_rows = [
        ("Baseline forecast (units_historical)", hist_vals),
        ("Enriched forecast (units_intent_adjusted)", enrich_vals),
        ("Δ vs baseline (units_delta)", [f"+{d}" if d > 0 else str(d) for d in delta_vals]),
        ("Actuals (units_actual)", actual_vals),
        ("Override (units_override)", ["—"] * len(WEEKS)),
    ]
    for name, vals in metric_rows:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=1).font = Font(bold=True)
        total = 0
        for j, v in enumerate(vals, 2):
            ws.cell(row=r, column=j, value=v)
            if isinstance(v, (int, float)):
                total += v
        if "Δ" not in name and "Actuals" not in name and "Override" not in name:
            ws.cell(row=r, column=len(WEEKS) + 2, value=total)
        elif name.startswith("Enriched"):
            ws.cell(row=r, column=len(WEEKS) + 2, value=sum(enrich_vals))
        elif name.startswith("Baseline"):
            ws.cell(row=r, column=len(WEEKS) + 2, value=sum(hist_vals))
        for c in range(1, len(headers_wide) + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="How to read this sheet")
    ws.cell(row=r, column=1).font = Font(bold=True)
    r += 1
    notes = [
        "Rows = metrics for ONE sku at a time (UI may list many SKUs with same column structure).",
        "Columns = one per week between From date and To date (here W42–W48 = 7 columns).",
        "Run 1: show Baseline + Actuals + Override only. Hide Enriched and Δ.",
        "Run 2: show all rows; Enriched replaces Baseline as primary number.",
        "Long format for pipelines → tab Sample_Long.",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note)
        r += 1

    auto_width(ws, 22)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
